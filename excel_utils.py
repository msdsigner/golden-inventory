"""
excel_utils.py
Utility functions for Golden Inventory automation.
"""

print("\n==============================")
print("USING excel_utils.py")
print(__file__)
print("==============================\n")

import re
from datetime import date, timedelta, datetime
from openpyxl import load_workbook
from scripts.item_overrides import ITEM_CATEGORY_OVERRIDES

# ─── Master 8-Pillar Taxonomy Rules ─────────────────────────────────────────
# ─── SKU PATTERN RULES ────────────────────────────────────────────────

SKU_PATTERN_RULES = [
    # Kitchen Appliances
    ("ED-", "Kitchen Appliances", "Heating & Cooking"),
    ("TS-", "Kitchen Appliances", "Air Fryer & Toaster"),
    ("FP", "Kitchen Appliances", "Food Preparation"),
    ("MG-", "Kitchen Appliances", "Food Preparation"),

    # Audio
    ("J-", "Audio", "Speakers"),
    ("BTSPK", "Audio", "Speakers"),
    ("PBX-", "Audio", "Speakers"),
    ("SP-", "Audio", "Speakers"),

    # TV / Mounts
    ("SM-STV", "Electronics", "TV Accessories"),
    ("SM-", "Electronics", "TV Accessories"),

    # Fans / Climate
    ("MPI-", "Air & Climate Control", "Fans"),
    ("LSL", "Air & Climate Control", "Fans"),

    # Coffee / Small Appliances
    ("CB", "Kitchen Appliances", "Coffee & Tea"),
]

MASTER_TAXONOMY_RULES = {

    "Major Appliances": {

        "Refrigerators": [
            "refrigerator", "fridge", "top freezer", "bottom freezer",
            "french door", "side by side", "counter depth"
        ],

        "Compact Refrigerators": [
            "compact refrigerator", "compact fridge",
            "mini fridge", "dorm fridge"
        ],

        "Beverage Centers": [
            "beverage center", "beverage cooler",
            "drink cooler"
        ],

        "Wine Coolers": [
            "wine cooler", "wine cellar"
        ],

        "Chest Freezers": [
            "chest freezer"
        ],

        "Upright Freezers": [
            "upright freezer",
            "convertible freezer"
        ],

        "Laundry": [
            "washer", "washing machine",
            "dryer", "stacked laundry",
            "laundry center"
        ],

        "Dishwashers": [
            "dishwasher"
        ],

        "Ranges": [
            "range", "gas range",
            "electric range"
        ],

        "Wall Ovens": [
            "wall oven"
        ],

        "Cooktops": [
            "cooktop", "induction cooktop"
        ],

        "Range Hoods": [
            "range hood", "hood"
        ]

    },

    "Kitchen Appliances": {

        "Coffee & Tea": [
            "coffee",
            "espresso",
            "keurig",
            "k-cup",
            "tea",
            "kettle",
            "coffee maker",
            "percolator",
            "urn",
            "milk frother"
        ],

        "Food Preparation": [
            "blender",
            "personal blender",
            "food processor",
            "processor",
            "chopper",
            "stand mixer",
            "hand mixer",
            "immersion blender",
            "hand blender",
            "juicer",
            "slow juicer",
            "cold press",
            "food chopper"
        ],

        "Air Fryer & Toaster": [
            "air fryer",
            "toaster",
            "toaster oven",
            "deep fryer"
        ],

        "Heating & Cooking": [
            "microwave",
            "hot plate",
            "hotplate",
            "burner",
            "double burner",
            "griddle",
            "pizza maker",
            "waffle",
            "waffle maker",
            "panini",
            "sandwich maker",
            "tortilla",
            "arepa",
            "dehydrator"
        ],

        "Pressure Cookers": [
            "pressure cooker",
            "pressure canner"
        ],

        "Rice Cookers": [
            "rice cooker"
        ],

        "Slow Cookers": [
            "slow cooker",
            "crock pot"
        ],

        "Cookware": [
            "cookware",
            "stock pot",
            "fry pan",
            "sauce pan",
            "skillet"
        ],

        "Kitchen Tools": [
            "knife sharpener",
            "can opener",
            "kitchen scale"
        ]

    },

    "Air & Climate Control": {

        "Portable AC": [
            "portable air conditioner",
            "portable ac"
        ],

        "Window AC": [
            "window air conditioner",
            "window ac"
        ],

        "Fans": [
            "fan",
            "tower fan",
            "pedestal fan",
            "desk fan",
            "neck fan",
            "box fan",
            "floor fan"
        ],

        "Air Purifiers": [
            "air purifier"
        ],

        "Humidifiers": [
            "humidifier"
        ],

        "Dehumidifiers": [
            "dehumidifier"
        ],

        "Heaters": [
            "heater",
            "ceramic heater",
            "radiator heater",
            "oil filled",
            "tower heater"
        ]

    },

    "Audio": {

        "Portable Audio": [
            "boombox",
            "cassette",
            "cd player",
            "portable cd",
            "walkman",
            "radio"
        ],

        "Home Audio": [
            "speaker",
            "bluetooth speaker",
            "soundbar",
            "subwoofer"
        ],

        "Headphones & Earphones": [
            "headphones",
            "earbuds",
            "earphones",
            "headset"
        ]

    },

    "Electronics": {

        "TV Accessories": [
            "antenna",
            "converter box"
        ],

        "TV Mounts": [
            "wall mount",
            "tv mount",
            "mount"
        ],

        "Remote Controls": [
            "remote",
            "universal remote"
        ],

        "Projectors": [
            "projector"
        ],

        "Media Players": [
            "media player",
            "dvd player",
            "streaming"
        ],

        "Security": [
            "security camera",
            "doorbell camera",
            "surveillance"
        ],

        "Car Electronics": [
            "car stereo",
            "fm transmitter"
        ],

        "Cables": [
            "hdmi",
            "usb cable",
            "audio cable"
        ],

        "Power Accessories": [
            "surge protector",
            "power strip",
            "extension cord"
        ]

    },

    "Health & Personal Care": {

        "Hair Care": [
            "hair dryer",
            "curling",
            "flat iron",
            "hot comb",
            "styler"
        ],

        "Personal Care": [
            "shaver",
            "trimmer",
            "clipper",
            "massager"
        ]

    },

    "Home & Office": {

        "Irons": [
            "iron",
            "steam iron"
        ],

        "Scales": [
            "scale",
            "body analysis"
        ],

        "Calculators": [
            "calculator"
        ],

        "Lighting": [
            "led light",
            "flashlight",
            "lamp"
        ],

        "Vacuums": [
            "vacuum"
        ],

        "Alarm Clocks": [
            "alarm clock"
        ]

    }

}


# ─── SKU NORMALIZATION ────────────────────────────────────────────────

def normalize_sku(sku: str) -> str:
    if not sku:
        return ""

    return (
        str(sku)
        .strip()
        .upper()
        .replace(" ", "")
        .replace("/", "-")
        .replace("--", "-")
    )


# ─── TAXONOMY ENGINE ────────────────────────────────────────────────

def get_taxonomy(product_category_path: str, item_id: str = None, description: str = ""):
    sku = str(item_id or "").upper()

    if sku in ITEM_CATEGORY_OVERRIDES:
        override = ITEM_CATEGORY_OVERRIDES[sku]
        # Overrides may be a full dict, a sub-category string, or other structure.
        if isinstance(override, dict):
            return override

        if isinstance(override, str):

            # If the override already matches one of our subcategories,
            # automatically determine its parent.
            for parent, subs in MASTER_TAXONOMY_RULES.items():
                if override in subs:
                    return {
                        "parent": parent,
                        "sub": override
                    }

            # Otherwise just preserve it.
            return {
                "parent": "Other",
                "sub": override
            }

        # Unknown override type: return it as-is
        return override

    is_refurbished = sku.endswith((
        "/RBO", "/RB", "/ROA",
        "-RBO", "-RB", "-ROA"
    ))

    test_str = (
        str(product_category_path) + " " +
        str(description) + " " +
        str(item_id)
    ).lower()

    # ---------- DEBUG (only for one SKU) ----------
    DEBUG_SKU = "58148G"

    if sku == DEBUG_SKU:
        print("\n========== TAXONOMY DEBUG ==========")
        print("SKU:", sku)
        print("Description:", description)
        print("Search String:", test_str)
        print("Kitchen Appliances exists:",
            "Kitchen Appliances" in MASTER_TAXONOMY_RULES)
        print("Food Preparation exists:",
            "Food Preparation" in MASTER_TAXONOMY_RULES.get("Kitchen Appliances", {}))
        print("====================================")

    # ---------- Smart Taxonomy Scoring ----------

    best_parent = None
    best_sub = None
    best_score = 0

    for parent, subs in MASTER_TAXONOMY_RULES.items():

        for sub, keywords in subs.items():

            score = 0

            for kw in keywords:

                kw = kw.lower()

                # Exact phrase match
                if kw in test_str:
                    score += len(kw.split()) * 20

                # Every individual word also contributes
                for word in kw.split():
                    if len(word) >= 3 and word in test_str:
                        score += 5

            if sku == DEBUG_SKU and score > 0:
                print(f"MATCH -> {parent} / {sub} = {score}")

            if score > best_score:
                best_score = score
                best_parent = parent
                best_sub = sub

    if sku == DEBUG_SKU:
        print("\n========== FINAL RESULT ==========")
        print("BEST PARENT:", best_parent)
        print("BEST SUB:", best_sub)
        print("BEST SCORE:", best_score)
        print("==================================\n")

    if best_parent:

        if is_refurbished:
            return {
                "parent": "Refurbished",
                "sub": best_sub
            }

        return {
            "parent": best_parent,
            "sub": best_sub
        }

    if is_refurbished:
        return {"parent": "Refurbished", "sub": "Uncategorized"}

    return {"parent": "Other", "sub": "Uncategorized"}

# ─── DATE HELPERS ────────────────────────────────────────────────

def parse_sheet_date(sheet_name: str) -> date:
    name = re.sub(r'\s*\(\d+\)\s*$', '', sheet_name).strip()

    for fmt in ["%B %d", "%B %d %Y"]:
        try:
            return datetime.strptime(name, fmt).date()
        except ValueError:
            continue

    return None

def get_most_recent_sheet_date(workbook):
    """
    Returns the date of the newest timestamped inventory sheet.
    Falls back to old weekly sheets only if no timestamped sheets exist.
    """

    latest_dt = None
    latest_weekly = None

    for sheet in workbook.sheetnames:

        if sheet == "Item Listing and Pricing":
            continue

        # New timestamped sheets
        try:
            dt = datetime.strptime(sheet, "%b %d %Y - %I-%M %p")

            if latest_dt is None or dt > latest_dt:
                latest_dt = dt

            continue

        except ValueError:
            pass

        # Old weekly sheets
        d = parse_sheet_date(sheet)

        if d and (latest_weekly is None or d > latest_weekly):
            latest_weekly = d

    if latest_dt:
        return latest_dt.date()

    return latest_weekly


def date_to_sheet_name(d: date) -> str:
    return d.strftime("%B %d").replace(" 0", " ")


def next_sheet_date(workbook) -> date:
    most_recent = get_most_recent_sheet_date(workbook)

    if most_recent:
        return most_recent + timedelta(weeks=1)

    today = date.today()
    return today + timedelta(days=(2 - today.weekday()) % 7)


# ─── RAW DATA LOADER ────────────────────────────────────────────────

def load_raw_data(inventory_tool_path: str) -> list[dict]:
    wb = load_workbook(inventory_tool_path, data_only=True)

    if "RAW" not in wb.sheetnames:
        raise ValueError("RAW sheet missing")

    ws = wb["RAW"]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() for h in rows[0]]

    data = []
    for row in rows[1:]:
        row_dict = dict(zip(headers, row))
        if any(v for v in row_dict.values() if v is not None):
            data.append(row_dict)

    wb.close()
    return data


# ─── CATEGORY WRAPPER ────────────────────────────────────────────────

def classify_category(product_category_path: str, item_id: str = None, description: str = "") -> str:
    return get_taxonomy(product_category_path, item_id, description)["sub"]

def get_latest_inventory_sheet(workbook):
    """
    Returns the newest timestamped inventory sheet and its date.
    """

    latest_sheet = None
    latest_datetime = None

    for sheet in workbook.sheetnames:

        if sheet == "Item Listing and Pricing":
            continue

        try:
            sheet_dt = datetime.strptime(sheet, "%b %d %Y - %I-%M %p")

            if latest_datetime is None or sheet_dt > latest_datetime:
                latest_datetime = sheet_dt
                latest_sheet = sheet

        except ValueError:
            continue

    if latest_sheet is None:
        raise Exception("No timestamped inventory sheet found.")

    return latest_sheet, latest_datetime.date()

def infer_parent_category(sub_category, text):

    mapping = {
        "Irons": "Home & Office",
        "Pressure Cooker": "Kitchen Appliances",
        "Coffee & Tea": "Kitchen Appliances",
        "Rice Cooker": "Kitchen Appliances",
        "Food Preparation": "Kitchen Appliances",
        "Slow Cookers": "Kitchen Appliances",
        "Electric Skillets": "Kitchen Appliances",

        "Wine Cooler": "Major Appliances",
        "Beverage Cooler": "Major Appliances",
        "Beverage Centers": "Major Appliances",
        "Compact Refrigerators": "Major Appliances",

        "Window AC": "Air & Climate Control",
        "Portable AC": "Air & Climate Control",

        "Speakers": "Audio",
        "Soundbars": "Audio",
        "Headphones & Earphones": "Audio",
        "Microphones": "Audio",

        "Phones": "Electronics & Communications",
        "Power Banks": "Electronics & Communications",
        "Computer Accessories": "Electronics & Communications",

        "Personal Care": "Health & Personal Care",

        "Laundry": "Laundry Appliances",
    }

    if sub_category in mapping:
        return mapping[sub_category]

    text = str(text).lower()

    if "iron" in text:
        return "Home & Office"

    if any(x in text for x in ["fridge", "cooler", "refrigerator"]):
        return "Major Appliances"

    if any(x in text for x in ["speaker", "radio", "microphone"]):
        return "Audio"

    if any(x in text for x in ["ac", "air conditioner", "fan"]):
        return "Air & Climate Control"

    return "Other"